#!/usr/bin/env python3
from __future__ import annotations

import argparse
import copy
import csv
import json
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Any


REPO_ROOT = Path(__file__).resolve().parents[1]
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

OPENPYXL_PATH = Path("/private/tmp/cats_excel_deps")
if str(OPENPYXL_PATH) not in sys.path:
    sys.path.insert(0, str(OPENPYXL_PATH))

from openpyxl import load_workbook  # type: ignore
from openpyxl.styles import PatternFill

from rag_eval.evaluator import aggregate_sample_results


ROOT = REPO_ROOT / "outputs" / "benchmark_local_committee_3judge"
LEGACY_CSV = ROOT / "master_results" / "cats_master_results_20260708.csv"
OUTPUT_XLSX = REPO_ROOT / "outputs" / "master_results_20260731_hierarchical.xlsx"
AUDIT_JSON = REPO_ROOT / "outputs" / "master_results_20260731_hierarchical_audit.json"

DATA_SHEET = "cats_master_results"
HEADER_ROW_1 = 1
HEADER_ROW_2 = 2
DATA_START_ROW = 3
DATA_END_ROW = 112
UPDATED_VALUE_COLUMNS = list(range(10, 28))  # J:AA

LEGACY_SIGNATURE_FIELDS = [
    "gr_answer_precision",
    "gr_answer_recall",
    "gr_answer_f1",
    "gr_refusal_precision",
    "gr_refusal_recall",
    "gr_refusal_f1",
    "gr_accuracy",
    "str",
    "fg",
    "behavior",
    "final_cats",
    "n",
    "str_n",
]
WORKBOOK_SIGNATURE_COLUMNS = [10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21, 24]


@dataclass(frozen=True)
class RowUpdate:
    excel_row: int
    source_relpath: str
    gr_answer_precision: float
    gr_answer_recall: float
    gr_answer_f1: float
    gr_refusal_precision: float
    gr_refusal_recall: float
    gr_refusal_f1: float
    gr_accuracy: float
    str_score: float
    fg: float
    behavior: float
    answer_quality: float
    cats_prevalence: float
    cats_balanced: float
    n: int
    behavior_n: int
    fg_n: int
    str_n: int
    answer_quality_n: int

    def to_sheet_values(self) -> list[float | int]:
        return [
            self.gr_answer_precision,
            self.gr_answer_recall,
            self.gr_answer_f1,
            self.gr_refusal_precision,
            self.gr_refusal_recall,
            self.gr_refusal_f1,
            self.gr_accuracy,
            self.str_score,
            self.fg,
            self.behavior,
            self.answer_quality,
            self.cats_prevalence,
            self.cats_balanced,
            self.n,
            self.behavior_n,
            self.fg_n,
            self.str_n,
            self.answer_quality_n,
        ]


def load_legacy_rows() -> list[dict[str, str]]:
    with LEGACY_CSV.open() as f:
        return list(csv.DictReader(f))


def workbook_signature(values: list[Any]) -> tuple[float, ...]:
    sig: list[float] = []
    for col in WORKBOOK_SIGNATURE_COLUMNS:
        value = values[col - 1]
        if value is None:
            raise ValueError(f"Missing signature cell in column {col}")
        sig.append(float(value))
    return tuple(sig)


def legacy_signature(row: dict[str, str]) -> tuple[float, ...]:
    return tuple(float(row[field]) for field in LEGACY_SIGNATURE_FIELDS)


def build_update_from_source(source_relpath: str) -> RowUpdate:
    path = ROOT / source_relpath
    data = json.loads(path.read_text())
    overall, _per_type, gr = aggregate_sample_results(data["per_sample"])
    return RowUpdate(
        excel_row=-1,
        source_relpath=source_relpath,
        gr_answer_precision=float(gr["precision"]),
        gr_answer_recall=float(gr["recall"]),
        gr_answer_f1=float(gr["f1"]),
        gr_refusal_precision=float(gr["abstain_precision"]),
        gr_refusal_recall=float(gr["abstain_recall"]),
        gr_refusal_f1=float(gr["abstain_f1"]),
        gr_accuracy=float(overall["gr_accuracy"]),
        str_score=float(overall["single_truth_recall"]),
        fg=float(overall["factual_grounding"]),
        behavior=float(overall["behavior"]),
        answer_quality=float(overall.get("answer_quality", 0.0)),
        cats_prevalence=float(overall["cats_prevalence_score"]),
        cats_balanced=float(overall["cats_balanced_score"]),
        n=int(overall["n"]),
        behavior_n=int(overall["behavior_n"]),
        fg_n=int(overall["factual_grounding_n"]),
        str_n=int(overall["single_truth_recall_n"]),
        answer_quality_n=int(overall.get("answer_quality_n", 0)),
    )


def extract_row_updates(workbook_path: Path) -> list[RowUpdate]:
    wb = load_workbook(workbook_path)
    ws = wb[DATA_SHEET]
    legacy_rows = load_legacy_rows()

    updates: list[RowUpdate] = []
    for row_idx in range(DATA_START_ROW, DATA_END_ROW + 1):
        values = [ws.cell(row_idx, col).value for col in range(1, 25)]
        if all(value is None for value in values):
            continue
        sig = workbook_signature(values)
        matches = []
        for legacy_row in legacy_rows:
            legacy_sig = legacy_signature(legacy_row)
            if all(abs(a - b) < 1e-12 for a, b in zip(sig, legacy_sig)):
                matches.append(legacy_row)
        if len(matches) != 1:
            raise ValueError(
                f"Could not match workbook row {row_idx} to a unique legacy result signature; found {len(matches)} matches"
            )
        matched_legacy = matches[0]
        row_update = build_update_from_source(matched_legacy["source_relpath"])
        # Preserve the legacy component cells exactly as stored in the supplied
        # workbook. The hierarchical CATS fields and applicability counts are
        # still recomputed from the latest source JSON above.
        updates.append(
            RowUpdate(
                excel_row=row_idx,
                source_relpath=row_update.source_relpath,
                gr_answer_precision=float(values[9]),
                gr_answer_recall=float(values[10]),
                gr_answer_f1=float(values[11]),
                gr_refusal_precision=float(values[12]),
                gr_refusal_recall=float(values[13]),
                gr_refusal_f1=float(values[14]),
                gr_accuracy=float(values[15]),
                str_score=float(values[16]),
                fg=float(values[17]),
                behavior=float(values[18]),
                answer_quality=row_update.answer_quality,
                cats_prevalence=row_update.cats_prevalence,
                cats_balanced=row_update.cats_balanced,
                n=row_update.n,
                behavior_n=row_update.behavior_n,
                fg_n=row_update.fg_n,
                str_n=row_update.str_n,
                answer_quality_n=row_update.answer_quality_n,
            )
        )
    if len(updates) != len(legacy_rows):
        raise ValueError(f"Expected {len(legacy_rows)} workbook rows, matched {len(updates)}")
    return updates


def set_column_widths(ws: Any) -> None:
    widths = {
        "Q": 32.13,
        "R": 30.38,
        "S": 34.13,
        "T": 22.0,
        "U": 22.0,
        "V": 20.0,
        "W": 13.0,
        "X": 13.0,
        "Y": 13.0,
        "Z": 13.0,
        "AA": 16.0,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def update_headers(ws: Any) -> None:
    # Extend separator-row merges before populating new columns.
    for merged in ("E99:X99", "E106:X106", "E99:AB99", "E106:AB106"):
        if merged in ws.merged_cells:
            ws.unmerge_cells(merged)
    ws.merge_cells("E99:AA99")
    ws.merge_cells("E106:AA106")

    # Extend the main header with vertically merged cells matching the existing style.
    for merged in ("Y1:Y2", "Z1:Z2", "AA1:AA2"):
        if merged not in ws.merged_cells:
            ws.merge_cells(merged)

    header_specs = {
        "Q1": "single_truth_recall",
        "R1": "factual_grounding",
        "S1": "behavioral_adherence",
        "T1": "answer_quality",
        "U1": "final_cats_prevalence",
        "V1": "final_cats_balanced",
        "W1": "n",
        "X1": "behavior_n",
        "Y1": "fg_n",
        "Z1": "str_n",
        "AA1": "answer_quality_n",
    }
    for cell_ref, value in header_specs.items():
        ws[cell_ref] = value

    # Preserve existing header styling.
    for target in ("Y1", "Z1", "AA1"):
        ws[target]._style = copy.copy(ws["X1"]._style)
        ws[target].fill = copy.copy(ws["X1"].fill)
    for target in ("Y2", "Z2", "AA2"):
        ws[target]._style = copy.copy(ws["X2"]._style)
        ws[target].fill = copy.copy(ws["X2"].fill)

    ws["AB1"] = None
    ws["AB2"] = None

    # Ensure the original populated header cells keep the same style while the text changes.
    for target in ("T1", "U1", "V1", "W1", "X1"):
        ws[target]._style = copy.copy(ws[target]._style)
    set_column_widths(ws)


def apply_updates(workbook_path: Path, output_path: Path, updates: list[RowUpdate]) -> None:
    wb = load_workbook(workbook_path)
    ws = wb[DATA_SHEET]
    update_headers(ws)

    for update in updates:
        row = update.excel_row
        values = update.to_sheet_values()
        # T:W should inherit the old final_cats style for visual consistency.
        old_final_style = copy.copy(ws.cell(row, 20)._style)
        count_style = copy.copy(ws.cell(row, 21)._style)

        for col, value in zip(UPDATED_VALUE_COLUMNS, values):
            ws.cell(row, col).value = value

        for col in (20, 21, 22):
            ws.cell(row, col)._style = copy.copy(old_final_style)
        for col in (23, 24, 25, 26, 27):
            ws.cell(row, col)._style = copy.copy(count_style)
        ws.cell(row, 28).value = None

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def verify_output(output_path: Path, updates: list[RowUpdate]) -> dict[str, Any]:
    wb = load_workbook(output_path, data_only=False)
    ws = wb[DATA_SHEET]

    expected_headers = {
        "Q1": "single_truth_recall",
        "R1": "factual_grounding",
        "S1": "behavioral_adherence",
        "T1": "answer_quality",
        "U1": "final_cats_prevalence",
        "V1": "final_cats_balanced",
        "W1": "n",
        "X1": "behavior_n",
        "Y1": "fg_n",
        "Z1": "str_n",
        "AA1": "answer_quality_n",
    }
    header_errors = {
        cell_ref: {"expected": expected, "actual": ws[cell_ref].value}
        for cell_ref, expected in expected_headers.items()
        if ws[cell_ref].value != expected
    }

    row_reports: list[dict[str, Any]] = []
    value_errors = 0
    for update in updates:
        row = update.excel_row
        actual = [ws.cell(row, col).value for col in UPDATED_VALUE_COLUMNS]
        expected = update.to_sheet_values()
        mismatches: list[dict[str, Any]] = []
        for col, act, exp in zip(UPDATED_VALUE_COLUMNS, actual, expected):
            if isinstance(exp, float):
                if abs(float(act) - exp) > 1e-12:
                    mismatches.append({"column": col, "expected": exp, "actual": act})
            else:
                if int(act) != exp:
                    mismatches.append({"column": col, "expected": exp, "actual": act})
        value_errors += len(mismatches)
        row_reports.append(
            {
                "excel_row": row,
                "source_relpath": update.source_relpath,
                "mismatch_count": len(mismatches),
                "mismatches": mismatches,
            }
        )

    return {
        "output_path": str(output_path),
        "updated_row_count": len(updates),
        "header_error_count": len(header_errors),
        "header_errors": header_errors,
        "value_error_count": value_errors,
        "rows": row_reports,
        "ok": not header_errors and value_errors == 0,
    }


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", type=Path, required=True)
    parser.add_argument("--output", type=Path, default=OUTPUT_XLSX)
    parser.add_argument("--audit-json", type=Path, default=AUDIT_JSON)
    args = parser.parse_args()

    updates = extract_row_updates(args.input)
    apply_updates(args.input, args.output, updates)
    audit = verify_output(args.output, updates)
    args.audit_json.parent.mkdir(parents=True, exist_ok=True)
    args.audit_json.write_text(json.dumps(audit, indent=2) + "\n")
    print(json.dumps(audit, indent=2))
    if not audit["ok"]:
        raise SystemExit(1)


if __name__ == "__main__":
    main()
